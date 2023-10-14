import datetime

import openpyxl
import paramiko

# pip install paramikoをインストールする必要があります

USERNAME = ""
PASSWORD = ""
RECV_SIZE = 1024
CODE = "ascii"
routers = ["", ]
IXs = ["", ]
command = "show configuration protocols bgp |display set | no-more"
ssh = paramiko.SSHClient()
excel_path = "./"

results = []

for router in routers:
    ssh.set_missing_host_key_policy(paramiko.MissingHostKeyPolicy())  # 未知の公開鍵の場合は無視
    ssh.connect(hostname=router, port=22, username=USERNAME, password=PASSWORD, timeout=60)  # 接続の確立

    stdin, stdout_conf, _ = ssh.exec_command(command)
    stdin.close()
    tmp_bgps = []
    tmp_description = ""
    tmp_asn = 0
    for line_stdout in stdout_conf:
        for ix in IXs:
            if "set protocols bgp group %s " % (ix,) in line_stdout:
                line_stdout_split = line_stdout.split(" ")
                if len(line_stdout_split) < 7:
                    continue
                now_ip = line_stdout_split[6].replace('\n', '')

                if len(line_stdout_split) == 9 and line_stdout_split[7] == "description":
                    tmp_description = line_stdout_split[8].replace('\n', '')
                elif len(line_stdout_split) == 9 and line_stdout_split[7] == "peer-as":
                    tmp_asn = line_stdout_split[8].replace('\n', '')
                if tmp_description != "" and tmp_asn != 0:
                    tmp_bgps.append({
                        "ip": now_ip,
                        "description": tmp_description,
                        "asn": tmp_asn
                    })
                    tmp_description = ""
                    tmp_asn = 0

    stdin, stdout_bgp_summary, stderr = ssh.exec_command("show bgp summary|no-more")
    for line_stdout in stdout_bgp_summary:
        for idx, tmp_bgp in enumerate(tmp_bgps):
            line_stdout_split = line_stdout.split()
            if line_stdout_split[0] == tmp_bgp["ip"] and line_stdout_split[1] == str(tmp_bgp["asn"]):
                tmp_bgps[idx]["status"] = line_stdout_split[len(line_stdout_split) - 1]

    stdin.close()

    results.append({
        "hostname": router,
        "result": stdout_conf,
        "bgp": tmp_bgps
    })

dt_now = datetime.datetime.now()
filename = dt_now.strftime('%Y-%m-%d_%H%M%S')
wb = openpyxl.Workbook()
for result in results:
    wb_sheet = wb.create_sheet(result["hostname"])
    for rows in range(2, len(result["bgp"]) + 1):
        bgp_row = result["bgp"][rows - 2]
        for cols in range(1, 6):
            cell = wb_sheet.cell(row=rows, column=cols)
            if cols == 1:
                cell.value = bgp_row["ip"]
            elif cols == 2:
                cell.value = bgp_row["description"]
            elif cols == 3:
                cell.value = bgp_row["asn"]
            elif cols == 4:
                cell.value = bgp_row.get("status", "")

wb.save(excel_path + "%s.xlsx" % filename)
